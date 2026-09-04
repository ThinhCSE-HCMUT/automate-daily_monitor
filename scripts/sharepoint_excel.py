#!/usr/bin/env python3
"""
Fill SharePoint Excel sheets (Voicelink / Fax / Virtual) from today's
daily_monitor.csv using Microsoft Graph Excel APIs (in-place cell updates).

Does not download/replace the whole .xlsx (avoids HTTP 423 file locks when
someone else has the workbook open). Dry-run still downloads locally.
"""
from __future__ import annotations

import argparse
import base64
import csv
import io
import json
import os
import sys
import time
from datetime import date, datetime, timedelta
from urllib.parse import quote, urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_TEXT

DATE_NUMBER_FORMAT = "m/d/yyyy"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CELL_ALIGN = Alignment(horizontal="center", vertical="center")
# Excel serial date epoch (Lotus 1900 bug compatible with Excel).
_EXCEL_EPOCH = date(1899, 12, 30)


SHEETS_FALLBACK = (
    {
        "imei": "866758040553188",
        "names": ("Voicelink Station No. 1", "Voicelink Station No.1"),
        "status_header": "Voicelink Status",
    },
    {
        "imei": "861107035967513",
        "names": ("Voicelink Station No. 2", "Voicelink Station No.2"),
        "status_header": "Voicelink Status",
    },
    {
        "imei": "861107035990853",
        "names": ("Fax Station No. 1", "Fax Station No.1"),
        "status_header": "Fax Status",
    },
    {
        "imei": "866758040526465",
        "names": ("Fax Station No. 2", "Fax Station No.2"),
        "status_header": "Fax Status",
    },
    {
        "imei": "866834045868010",
        "names": ("Virtual Station No. 1", "Virtual Station No.1"),
        "status_header": "",
    },
    {
        "imei": "866834041157558",
        "names": ("Virtual Station No. 2", "Virtual Station No.2"),
        "status_header": "",
    },
)


def resolve_sheets(monitor_conf: str):
    try:
        from stations_lib import load_stations, sheets_specs

        specs = sheets_specs(load_stations(monitor_conf))
        if specs:
            return specs
    except Exception as exc:
        log("WARN", f"Could not load SharePoint sheets from {monitor_conf}: {exc}")
    return list(SHEETS_FALLBACK)

def log(level: str, msg: str) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{now}] [{level}] {msg}"
    try:
        print(line, flush=True)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(line.encode(enc, errors="replace").decode(enc, errors="replace"), flush=True)


def load_conf(path: str) -> dict[str, str]:
    cfg: dict[str, str] = {}
    if not os.path.isfile(path):
        return cfg
    with open(path, encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            cfg[key.strip()] = val.strip()
    return cfg


def norm_header(s: str) -> str:
    return " ".join((s or "").replace("\n", " ").split()).strip().lower()


def parse_day(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip().strip('"')
    if not s:
        return None
    for fmt, width in (
        ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y-%m-%d", 10),
        ("%m/%d/%Y", 10),
        ("%d/%m/%Y", 10),
    ):
        try:
            return datetime.strptime(s[:width], fmt).date()
        except ValueError:
            continue
    return None


def encode_share_id(url: str) -> str:
    b64 = base64.b64encode(url.encode("utf-8")).decode("ascii")
    return "u!" + b64.rstrip("=").replace("/", "_").replace("+", "-")


def read_today_csv(path: str, today: date) -> dict[str, dict[str, str]]:
    """Rows whose Date is `today` (not the whole CSV). Last row per IMEI wins."""
    rows: dict[str, dict[str, str]] = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return rows
        for raw in reader:
            rec = {k.strip(): (v or "").strip() for k, v in raw.items() if k}
            imei = rec.get("IMEI") or ""
            day = parse_day(rec.get("Date") or "")
            if not imei or day != today:
                continue
            rows[imei] = rec
    return rows


# All station sheets: "Uptime (days, hh:mm)". CSV still uses "Uptime (hh:mm)".
EXCEL_UPTIME = "Uptime (days, hh:mm)"
HEADER_ALIASES = {
    "uptime (days, hh:mm)": (
        "Uptime (days, hh:mm)",
        "Uptime (day:hh:mm)",
        "Uptime (hh:mm)",
        "Uptime",
    ),
    "uptime (hh:mm)": (
        "Uptime (days, hh:mm)",
        "Uptime (day:hh:mm)",
        "Uptime (hh:mm)",
        "Uptime",
    ),
    "wifi status (sim data)": (
        "WiFi Status (Sim Data)",
        "WiFi Status",
    ),
    "voicelink/fax status": ("Voicelink/Fax status", "Voicelink Status", "Fax Status"),
    "fax status": ("Fax Status", "Voicelink/Fax status"),
    "voicelink status": ("Voicelink Status", "Voicelink/Fax status"),
}


def rec_field(rec: dict[str, str], *names: str) -> str:
    by_norm = {norm_header(k): (v or "").strip() for k, v in rec.items() if k}
    for name in names:
        v = by_norm.get(norm_header(name))
        if v:
            return v
    for name in names:
        token = norm_header(name).split()[0]
        if not token:
            continue
        for key, val in by_norm.items():
            if key.startswith(token) and val:
                return val
    return ""


def csv_to_excel_values(rec: dict[str, str], status_header: str) -> dict[str, str]:
    values = {
        "Date": rec_field(rec, "Date"),
        "Anydesk ID": rec_field(rec, "Anydesk ID"),
        "IMEI": rec_field(rec, "IMEI"),
        "Firmware Version": rec_field(rec, "Firmware Version"),
        EXCEL_UPTIME: rec_field(rec, "Uptime (hh:mm)", EXCEL_UPTIME, "Uptime (day:hh:mm)", "Uptime"),
        "Carrier": rec_field(rec, "Carrier"),
        "Phone": rec_field(rec, "Phone"),
        "RSSI (dBm)": rec_field(rec, "RSSI (dBm)", "RSSI"),
        "WiFi Status (Sim Data)": rec_field(rec, "WiFi Status (Sim Data)", "WiFi Status"),
        "SSH Access": rec_field(rec, "SSH Access"),
        "Note": rec_field(rec, "Note"),
    }
    if status_header == "Voicelink Status":
        values[status_header] = "N/A"
    elif status_header:
        values[status_header] = rec_field(
            rec, "Voicelink/Fax status", "Voicelink Status", "Fax Status"
        )
    return values


def find_header_row(ws, max_scan: int = 12) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_map: dict[str, int] = {}
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        mapping: dict[str, int] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            mapping[norm_header(str(val))] = c
        if "imei" in mapping and "date" in mapping and len(mapping) > len(best_map):
            best_row, best_map = r, mapping
    return best_row, best_map


def col_for(mapping: dict[str, int], *names: str) -> int | None:
    wanted: list[str] = []
    for name in names:
        wanted.append(norm_header(name))
        for alias in HEADER_ALIASES.get(norm_header(name), ()):
            wanted.append(norm_header(alias))
    seen: set[str] = set()
    for key in wanted:
        if not key or key in seen:
            continue
        seen.add(key)
        idx = mapping.get(key)
        if idx:
            return idx
    for name in names:
        token = norm_header(name).split()[0]
        if not token:
            continue
        for key, idx in mapping.items():
            if key.startswith(token):
                return idx
    return None


def last_used_row(ws, header_row: int, key_col: int) -> int:
    last = header_row
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if ws.cell(r, key_col).value not in (None, ""):
            last = r
    return last


def find_same_day_row(ws, header_row: int, date_col: int, today: date) -> int | None:
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if parse_day(ws.cell(r, date_col).value) == today:
            return r
    return None


def expand_tables(ws, max_row: int) -> None:
    for tbl in list(getattr(ws, "tables", {}).values()):
        try:
            min_col, min_row, max_col, old_max = range_boundaries(tbl.ref)
        except Exception:
            continue
        if max_row > old_max:
            tbl.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}"
            )


def write_date_cell(ws, row: int, col: int, value, fallback: date) -> None:
    day = parse_day(value) or fallback
    cell = ws.cell(row, col)
    cell.value = datetime(day.year, day.month, day.day)
    cell.number_format = DATE_NUMBER_FORMAT


def apply_row_style(ws, row: int, start_col: int, end_col: int) -> None:
    if start_col > end_col:
        start_col, end_col = end_col, start_col
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.border = THIN_BORDER
        cell.alignment = CELL_ALIGN


def write_cell(ws, row: int, col: int, value, force_text: bool) -> None:
    cell = ws.cell(row, col)
    if force_text:
        cell.number_format = FORMAT_TEXT
        cell.value = "" if value is None else str(value)
        return
    if isinstance(cell.value, datetime) and isinstance(value, str):
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(value[:19] if " " in fmt else value[:10], fmt)
                break
            except ValueError:
                continue
        cell.value = parsed if parsed else value
        return
    cell.value = value


def resolve_sheet(wb, names: tuple[str, ...]):
    exact = {n.lower(): n for n in wb.sheetnames}
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
        hit = exact.get(name.lower())
        if hit:
            return wb[hit]
    for sheet_name in wb.sheetnames:
        low = sheet_name.lower()
        if any(n.lower() in low or low in n.lower() for n in names):
            return wb[sheet_name]
    return None


def upsert_sheet(wb, spec: dict, rec: dict[str, str], today: date) -> str:
    ws = resolve_sheet(wb, spec["names"])
    if ws is None:
        return f"sheet not found (tried {', '.join(spec['names'])})"

    header_row, mapping = find_header_row(ws)
    date_col = col_for(mapping, "Date")
    imei_col = col_for(mapping, "IMEI")
    if not date_col or not imei_col:
        return f"{ws.title}: missing Date/IMEI header"

    values = csv_to_excel_values(rec, spec["status_header"])
    uptime_val = values.get(EXCEL_UPTIME) or ""
    if not uptime_val or uptime_val.upper() in ("N/A", "NA"):
        log(
            "WARN",
            f"{ws.title}: CSV uptime is {uptime_val!r} for IMEI {rec.get('IMEI') or spec['imei']}",
        )
    existing = find_same_day_row(ws, header_row, date_col, today)
    if existing:
        target = existing
        action = "fill-blank"
        only_blank = True
    else:
        target = last_used_row(ws, header_row, imei_col) + 1
        action = "append"
        only_blank = False

    used_cols: list[int] = []
    wrote = 0
    for header, value in values.items():
        col = col_for(mapping, header)
        if not col:
            if "uptime" in header.lower():
                log(
                    "WARN",
                    f"{ws.title}: no Uptime column (headers={sorted(mapping.keys())})",
                )
            continue
        used_cols.append(col)
        if only_blank:
            cur = ws.cell(target, col).value
            if cur not in (None, ""):
                continue
            if value is None or str(value).strip() == "":
                continue
        if header == "Date":
            if not only_blank:
                write_date_cell(ws, target, col, value, today)
                wrote += 1
            continue
        force_text = header in ("IMEI", "Anydesk ID", "Phone")
        write_cell(ws, target, col, value, force_text)
        wrote += 1
    if only_blank and wrote == 0:
        return f"{ws.title}: already has {today.isoformat()} (complete)"
    note_col = col_for(mapping, "Note")
    end_col = note_col or (max(used_cols) if used_cols else date_col)
    apply_row_style(ws, target, date_col, end_col)
    expand_tables(ws, target)
    if only_blank:
        return f"{ws.title} row {target} (filled {wrote} blank cells)"
    return f"{ws.title} row {target} ({action})"


GRAPH_SCOPES = (
    "https://graph.microsoft.com/Files.ReadWrite.All",
    "https://graph.microsoft.com/Sites.ReadWrite.All",
    "https://graph.microsoft.com/User.Read",
)
# Microsoft Office is already trusted (same family as Excel on Edge). Do NOT use
# Graph Command Line Tools — that app needs admin consent in this tenant.
OFFICE_CLIENT_ID = "d3590ed6-52b3-4102-aeff-aad2292ab01c"
AZURE_CLI_CLIENT_ID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"
DEFAULT_CLIENT_IDS = (OFFICE_CLIENT_ID, AZURE_CLI_CLIENT_ID)


def sharepoint_origin(site_url: str) -> str:
    p = urlparse(site_url)
    return f"{p.scheme}://{p.netloc}"


def spo_scope_sets(site_url: str) -> list[list[str]]:
    origin = sharepoint_origin(site_url)
    return [
        [f"{origin}/.default"],
        [f"{origin}/AllSites.Write"],
        ["https://graph.microsoft.com/.default"],
        list(GRAPH_SCOPES),
    ]


def token_kind_from_scopes(scopes: list[str]) -> str:
    joined = " ".join(scopes).lower()
    if "graph.microsoft.com" in joined:
        return "graph"
    return "spo"


def infer_tenant(username: str, cfg_tenant: str) -> str:
    if cfg_tenant:
        return cfg_tenant
    if "@" in username:
        return username.split("@", 1)[1].strip()
    return "simplifi.io"


def client_ids_to_try(cfg_id: str) -> list[str]:
    ids: list[str] = []
    if cfg_id:
        ids.append(cfg_id)
    for cid in DEFAULT_CLIENT_IDS:
        if cid not in ids:
            ids.append(cid)
    return ids


def _save_msal_cache(cache, path: str) -> None:
    if cache.has_state_changed:
        with open(path, "w", encoding="utf-8") as f:
            f.write(cache.serialize())


def edge_exe() -> str | None:
    if sys.platform != "win32":
        return None
    for path in (
        os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
        os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
    ):
        if os.path.isfile(path):
            return path
    return None


def open_edge(url: str) -> None:
    import subprocess
    import webbrowser

    exe = edge_exe()
    if exe:
        subprocess.Popen([exe, url])
        return
    webbrowser.open(url)


def _msal_app(tenant: str, client_id: str, cache):
    import msal

    # Do not enable the Windows broker here. In Cursor/VS Code the WAM dialog
    # stays behind the editor and acquire_token_interactive waits forever.
    return msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant}",
        token_cache=cache,
    )


def interactive_attempts(site_url: str, cfg_client_id: str) -> list[tuple[str, list[str]]]:
    origin = sharepoint_origin(site_url)
    attempts: list[tuple[str, list[str]]] = []
    if cfg_client_id:
        attempts.append((cfg_client_id, [f"{origin}/.default"]))
    attempts.extend(
        [
            (OFFICE_CLIENT_ID, [f"{origin}/.default"]),
            (OFFICE_CLIENT_ID, ["https://graph.microsoft.com/.default"]),
            (AZURE_CLI_CLIENT_ID, [f"{origin}/.default"]),
        ]
    )
    seen: set[tuple[str, tuple[str, ...]]] = set()
    out: list[tuple[str, list[str]]] = []
    for cid, scopes in attempts:
        key = (cid, tuple(scopes))
        if key not in seen:
            seen.add(key)
            out.append((cid, scopes))
    return out


def acquire_token_silent_or_password(
    tenant: str,
    client_id: str,
    username: str,
    password: str,
    cache_path: str,
    scopes: list[str],
    try_password: bool,
) -> str | None:
    import msal

    cache = msal.SerializableTokenCache()
    if os.path.isfile(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            cache.deserialize(f.read())
    app = _msal_app(tenant, client_id, cache)
    accounts = app.get_accounts(username=username) if username else app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
        if result and result.get("access_token"):
            _save_msal_cache(cache, cache_path)
            return result["access_token"]
    if try_password and password:
        result = app.acquire_token_by_username_password(username, password, scopes=scopes)
        _save_msal_cache(cache, cache_path)
        if result and result.get("access_token"):
            return result["access_token"]
        err = result.get("error_description") or result.get("error") or ""
        raise RuntimeError(err)
    return None


def acquire_token_device_code(
    tenant: str, client_id: str, cache_path: str, scopes: list[str]
) -> str:
    import msal

    cache = msal.SerializableTokenCache()
    if os.path.isfile(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            cache.deserialize(f.read())
    app = _msal_app(tenant, client_id, cache)
    flow = app.initiate_device_flow(scopes=scopes)
    if "user_code" not in flow:
        raise RuntimeError(flow.get("error_description") or flow.get("error") or str(flow))
    uri = flow.get("verification_uri_complete") or flow.get("verification_uri") or "https://microsoft.com/devicelogin"
    mins = max(1, int(flow.get("expires_in", 900)) // 60)
    log("INFO", "Not frozen — waiting for device-code sign-in in Edge")
    log("INFO", f"1. Open {uri}  (Edge should pop up)")
    log("INFO", f"2. Enter code: {flow['user_code']}")
    log("INFO", "3. Pick thinh.le.aura@simplifi.io")
    log("INFO", f"Waiting up to {mins} min. Ctrl+C to abort.")
    open_edge(uri)
    result = app.acquire_token_by_device_flow(flow)
    _save_msal_cache(cache, cache_path)
    if result and result.get("access_token"):
        log("INFO", f"Device-code login OK (client {client_id[:8]}…)")
        return result["access_token"]
    raise RuntimeError(result.get("error_description") or result.get("error") or "device-code login failed")


def msal_login(
    tenant: str,
    username: str,
    password: str,
    cfg_client_id: str,
    cache_path: str,
    site_url: str,
    allow_interactive: bool,
) -> tuple[str, str]:
    """Return (access_token, 'graph'|'spo'). Prefer Office + SharePoint (no Graph CLI consent)."""
    try_password = bool(password) and sys.platform != "win32"
    last: Exception | None = None
    for cid in client_ids_to_try(cfg_client_id):
        for scopes in spo_scope_sets(site_url):
            kind = token_kind_from_scopes(scopes)
            try:
                token = acquire_token_silent_or_password(
                    tenant, cid, username, password, cache_path, scopes, try_password
                )
                if token:
                    log("INFO", f"Token from cache/ROPC ({kind}, client {cid[:8]}…)")
                    return token, kind
            except Exception as exc:
                last = exc
                log("DEBUG", f"Silent/ROPC {cid[:8]} {kind}: {exc}")
    if allow_interactive:
        for cid, scopes in interactive_attempts(site_url, cfg_client_id):
            kind = token_kind_from_scopes(scopes)
            try:
                token = acquire_token_device_code(tenant, cid, cache_path, scopes)
                return token, kind
            except Exception as exc:
                last = exc
                msg = str(exc).lower()
                log("DEBUG", f"Device code {cid[:8]} {kind}: {exc}")
                if "cancel" in msg or "user_canceled" in msg or "access_denied" in msg:
                    raise RuntimeError("Sign-in was cancelled") from exc
    raise RuntimeError(str(last) if last else "Microsoft login failed")


def api_root_for(kind: str, site_url: str) -> str:
    if kind == "graph":
        return "https://graph.microsoft.com/v1.0"
    return f"{sharepoint_origin(site_url)}/_api/v2.0"


def api_request(token: str, method: str, url: str, **kwargs):
    import requests

    headers = dict(kwargs.pop("headers", {}) or {})
    headers["Authorization"] = f"Bearer {token}"
    headers.setdefault("Accept", "application/json")
    resp = requests.request(method, url, headers=headers, timeout=120, **kwargs)
    if resp.status_code >= 400:
        raise RuntimeError(f"{method} {url} HTTP {resp.status_code}: {resp.text[:500]}")
    return resp


def resolve_share(token: str, file_url: str, api_root: str) -> tuple[str, str] | None:
    urls = [file_url]
    if "?" in file_url:
        urls.append(file_url.split("?", 1)[0])
    for raw in urls:
        share_id = encode_share_id(raw)
        url = f"{api_root}/shares/{share_id}/driveItem"
        try:
            payload = api_request(token, "GET", url).json()
            item_id = payload.get("id")
            drive_id = (payload.get("parentReference") or {}).get("driveId")
            name = payload.get("name") or ""
            if item_id and drive_id:
                log("INFO", f"Resolved sharing link -> {name}")
                return drive_id, item_id
        except Exception as exc:
            log("DEBUG", f"shares {share_id[:24]}…: {exc}")
    return None


def download_drive_item(token: str, drive_id: str, item_id: str, api_root: str) -> bytes:
    url = f"{api_root}/drives/{drive_id}/items/{item_id}/content"
    data = api_request(token, "GET", url).content
    if not data:
        raise RuntimeError("Empty workbook download")
    return data


def upload_drive_item(token: str, drive_id: str, item_id: str, api_root: str, data: bytes) -> None:
    url = f"{api_root}/drives/{drive_id}/items/{item_id}/content"
    api_request(
        token,
        "PUT",
        url,
        data=data,
        headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )


def excel_serial(day: date) -> float:
    return float((day - _EXCEL_EPOCH).days)


def graph_excel_base(drive_id: str, item_id: str) -> str:
    return f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"


def graph_excel_request(
    token: str,
    method: str,
    url: str,
    session_id: str | None = None,
    **kwargs,
):
    import requests

    headers = dict(kwargs.pop("headers", {}) or {})
    headers["Authorization"] = f"Bearer {token}"
    headers.setdefault("Accept", "application/json")
    if session_id:
        headers["Workbook-Session-Id"] = session_id
    if "json" in kwargs and "Content-Type" not in headers:
        headers["Content-Type"] = "application/json"
    resp = requests.request(method, url, headers=headers, timeout=120, **kwargs)
    if resp.status_code >= 400:
        raise RuntimeError(f"{method} {url} HTTP {resp.status_code}: {resp.text[:500]}")
    if resp.status_code == 204 or not resp.content:
        return None
    try:
        return resp.json()
    except Exception:
        return None


def create_workbook_session(token: str, drive_id: str, item_id: str, persist: bool = True) -> str:
    url = f"{graph_excel_base(drive_id, item_id)}/createSession"
    payload = graph_excel_request(
        token, "POST", url, json={"persistChanges": persist}
    ) or {}
    sid = payload.get("id")
    if not sid:
        raise RuntimeError(f"createSession missing id: {payload}")
    return sid


def close_workbook_session(token: str, drive_id: str, item_id: str, session_id: str) -> None:
    url = f"{graph_excel_base(drive_id, item_id)}/closeSession"
    try:
        graph_excel_request(token, "POST", url, session_id=session_id, json={})
    except Exception as exc:
        log("DEBUG", f"closeSession: {exc}")


def list_worksheet_names(token: str, drive_id: str, item_id: str, session_id: str) -> list[str]:
    url = f"{graph_excel_base(drive_id, item_id)}/worksheets?$select=name,position"
    payload = graph_excel_request(token, "GET", url, session_id=session_id) or {}
    items = payload.get("value") or []
    items = sorted(items, key=lambda x: x.get("position", 0))
    return [str(x.get("name") or "") for x in items if x.get("name")]


def resolve_sheet_name(names_available: list[str], wanted: tuple[str, ...]) -> str | None:
    exact = {n.lower(): n for n in names_available}
    for name in wanted:
        if name in names_available:
            return name
        hit = exact.get(name.lower())
        if hit:
            return hit
    for sheet_name in names_available:
        low = sheet_name.lower()
        if any(n.lower() in low or low in n.lower() for n in wanted):
            return sheet_name
    return None


def worksheet_used_range(
    token: str, drive_id: str, item_id: str, session_id: str, sheet_name: str
) -> tuple[list[list], int, int]:
    """Return (values matrix 0-index, row_count, column_count)."""
    enc = quote(sheet_name, safe="")
    url = (
        f"{graph_excel_base(drive_id, item_id)}/worksheets('{enc}')/"
        f"usedRange(valuesOnly=true)?$select=values,rowCount,columnCount,address"
    )
    payload = graph_excel_request(token, "GET", url, session_id=session_id) or {}
    values = payload.get("values") or []
    rows = int(payload.get("rowCount") or len(values) or 0)
    cols = int(payload.get("columnCount") or (len(values[0]) if values else 0))
    # Normalize ragged rows
    matrix: list[list] = []
    for r in range(rows):
        row = list(values[r]) if r < len(values) else []
        if len(row) < cols:
            row.extend([None] * (cols - len(row)))
        matrix.append(row[:cols] if cols else row)
    return matrix, rows, cols


def find_header_row_matrix(matrix: list[list], max_scan: int = 12) -> tuple[int, dict[str, int]]:
    """Return 0-based header row index and mapping norm_header -> 0-based col."""
    best_i = 0
    best_map: dict[str, int] = {}
    scan = min(max_scan, len(matrix))
    for i in range(scan):
        mapping: dict[str, int] = {}
        for c, val in enumerate(matrix[i]):
            if val is None or val == "":
                continue
            mapping[norm_header(str(val))] = c
        if "imei" in mapping and "date" in mapping and len(mapping) > len(best_map):
            best_i, best_map = i, mapping
    return best_i, best_map


def cell_empty(val) -> bool:
    return val is None or val == ""


def parse_day_cell(value) -> date | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return _EXCEL_EPOCH + timedelta(days=int(value))
        except Exception:
            return None
    return parse_day(value)


def patch_cells(
    token: str,
    drive_id: str,
    item_id: str,
    session_id: str,
    sheet_name: str,
    updates: list[tuple[int, int, object, str | None]],
) -> int:
    """
    updates: list of (row_1based, col_1based, value, number_format_or_None).
    Batches contiguous same-row updates where possible.
    """
    if not updates:
        return 0
    wrote = 0
    by_row: dict[int, list[tuple[int, object, str | None]]] = {}
    for r, c, v, fmt in updates:
        by_row.setdefault(r, []).append((c, v, fmt))
    for row, cells in by_row.items():
        cells.sort(key=lambda x: x[0])
        run: list[tuple[int, object, str | None]] = []

        def flush() -> None:
            nonlocal wrote, run
            if not run:
                return
            c0, c1 = run[0][0], run[-1][0]
            addr = f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}"
            vals = [[x[1] for x in run]]
            fmts = [[x[2] or "General" for x in run]]
            need_fmt = any(x[2] for x in run)
            enc = quote(sheet_name, safe="")
            addr_q = quote(addr, safe="")
            url = (
                f"{graph_excel_base(drive_id, item_id)}/worksheets('{enc}')/"
                f"range(address='{addr_q}')"
            )
            body: dict = {"values": vals}
            if need_fmt:
                body["numberFormat"] = fmts
            graph_excel_request(token, "PATCH", url, session_id=session_id, json=body)
            wrote += len(run)
            run = []

        prev = None
        for c, v, fmt in cells:
            if prev is not None and c != prev + 1:
                flush()
            run.append((c, v, fmt))
            prev = c
        flush()
    return wrote


def ensure_table_covers_row(
    token: str,
    drive_id: str,
    item_id: str,
    session_id: str,
    sheet_name: str,
    target_row_1based: int,
) -> None:
    """Expand worksheet tables so they include target_row if needed."""
    enc = quote(sheet_name, safe="")
    url = f"{graph_excel_base(drive_id, item_id)}/worksheets('{enc}')/tables?$select=name,id"
    try:
        payload = graph_excel_request(token, "GET", url, session_id=session_id) or {}
    except Exception as exc:
        log("DEBUG", f"list tables: {exc}")
        return
    for tbl in payload.get("value") or []:
        tid = tbl.get("id") or tbl.get("name")
        if not tid:
            continue
        try:
            turl = (
                f"{graph_excel_base(drive_id, item_id)}/worksheets('{enc}')/"
                f"tables('{quote(str(tid), safe='')}')/range?$select=address"
            )
            rng = graph_excel_request(token, "GET", turl, session_id=session_id) or {}
            address = rng.get("address") or ""
            # address like "'Sheet'!A1:L20" or "A1:L20"
            if "!" in address:
                address = address.split("!", 1)[1]
            address = address.replace("$", "")
            if ":" not in address:
                continue
            start, end = address.split(":", 1)
            # parse end row digits
            end_row = int("".join(ch for ch in end if ch.isdigit()) or "0")
            start_col = "".join(ch for ch in start if ch.isalpha())
            end_col = "".join(ch for ch in end if ch.isalpha())
            start_row = int("".join(ch for ch in start if ch.isdigit()) or "1")
            if target_row_1based <= end_row:
                continue
            new_addr = f"{start_col}{start_row}:{end_col}{target_row_1based}"
            # resize via DataBodyRange is awkward; use table resize endpoint if available
            resize_url = (
                f"{graph_excel_base(drive_id, item_id)}/worksheets('{enc}')/"
                f"tables('{quote(str(tid), safe='')}')/resize"
            )
            # Some tenants use POST resize with newRange; fall back to range PATCH on worksheet
            try:
                graph_excel_request(
                    token,
                    "POST",
                    resize_url,
                    session_id=session_id,
                    json={"newRange": new_addr},
                )
                log("DEBUG", f"Resized table {tid} -> {new_addr}")
            except Exception as exc:
                log("DEBUG", f"table resize {tid}: {exc}")
        except Exception as exc:
            log("DEBUG", f"table cover: {exc}")


def upsert_sheet_graph(
    token: str,
    drive_id: str,
    item_id: str,
    session_id: str,
    sheet_names: list[str],
    spec: dict,
    rec: dict[str, str],
    today: date,
) -> str:
    title = resolve_sheet_name(sheet_names, spec["names"])
    if not title:
        return f"sheet not found (tried {', '.join(spec['names'])})"

    matrix, _rows, _cols = worksheet_used_range(token, drive_id, item_id, session_id, title)
    if not matrix:
        return f"{title}: empty sheet"
    header_i, mapping = find_header_row_matrix(matrix)
    date_c = col_for(mapping, "Date")
    imei_c = col_for(mapping, "IMEI")
    if date_c is None or imei_c is None:
        return f"{title}: missing Date/IMEI header"
    # mapping values are 0-based; col_for returns those indices

    values = csv_to_excel_values(rec, spec["status_header"])
    uptime_val = values.get(EXCEL_UPTIME) or ""
    if not uptime_val or uptime_val.upper() in ("N/A", "NA"):
        log(
            "WARN",
            f"{title}: CSV uptime is {uptime_val!r} for IMEI {rec.get('IMEI') or spec['imei']}",
        )

    # Find same-day row (0-based in matrix)
    existing_i: int | None = None
    for r in range(header_i + 1, len(matrix)):
        if parse_day_cell(matrix[r][date_c] if date_c < len(matrix[r]) else None) == today:
            existing_i = r
            break

    if existing_i is not None:
        target_i = existing_i
        only_blank = True
        action = "fill-blank"
    else:
        last = header_i
        for r in range(header_i + 1, len(matrix)):
            cell = matrix[r][imei_c] if imei_c < len(matrix[r]) else None
            if not cell_empty(cell):
                last = r
        target_i = last + 1
        only_blank = False
        action = "append"

    target_row_1 = target_i + 1  # Graph/Excel 1-based
    updates: list[tuple[int, int, object, str | None]] = []

    for header, value in values.items():
        c0 = col_for(mapping, header)
        if c0 is None:
            if "uptime" in header.lower():
                log(
                    "WARN",
                    f"{title}: no Uptime column (headers={sorted(mapping.keys())})",
                )
            continue
        col_1 = c0 + 1
        if only_blank:
            cur = None
            if target_i < len(matrix) and c0 < len(matrix[target_i]):
                cur = matrix[target_i][c0]
            if not cell_empty(cur):
                continue
            if value is None or str(value).strip() == "":
                continue
        if header == "Date":
            if only_blank:
                continue
            day = parse_day(value) or today
            updates.append((target_row_1, col_1, excel_serial(day), DATE_NUMBER_FORMAT))
            continue
        force_text = header in ("IMEI", "Anydesk ID", "Phone")
        cell_val: object = "" if value is None else str(value)
        fmt = "@" if force_text else None
        updates.append((target_row_1, col_1, cell_val, fmt))

    if only_blank and not updates:
        return f"{title}: already has {today.isoformat()} (complete)"

    wrote = patch_cells(token, drive_id, item_id, session_id, title, updates)
    if action == "append":
        ensure_table_covers_row(token, drive_id, item_id, session_id, title, target_row_1)

    if only_blank:
        return f"{title} row {target_row_1} (filled {wrote} blank cells)"
    return f"{title} row {target_row_1} ({action})"


def fill_workbook_via_graph(
    token: str,
    drive_id: str,
    item_id: str,
    csv_rows: dict[str, dict[str, str]],
    today: date,
    sheets: list | None = None,
) -> int:
    """In-place Excel updates via Graph. Returns 0 on success."""
    session_id = create_workbook_session(token, drive_id, item_id, persist=True)
    log("INFO", "Opened Graph Excel workbook session (in-place edit)")
    try:
        names = list_worksheet_names(token, drive_id, item_id, session_id)
        log("INFO", f"Workbook sheets: {', '.join(names)}")
        ok = 0
        changed = 0
        for spec in sheets or SHEETS_FALLBACK:
            rec = csv_rows.get(spec["imei"])
            if not rec:
                log(
                    "WARN",
                    f"No CSV row today for IMEI {spec['imei']} ({spec['names'][0]}) — skip sheet",
                )
                continue
            # Retry once on transient lock/conflict
            last_exc: Exception | None = None
            for attempt in range(3):
                try:
                    result = upsert_sheet_graph(
                        token, drive_id, item_id, session_id, names, spec, rec, today
                    )
                    last_exc = None
                    break
                except Exception as exc:
                    last_exc = exc
                    msg = str(exc).lower()
                    if attempt < 2 and ("423" in msg or "locked" in msg or "conflict" in msg):
                        time.sleep(2 + attempt * 2)
                        continue
                    raise
            if last_exc:
                raise last_exc
            if result.startswith("sheet not found") or "missing" in result:
                log("ERROR", result)
                continue
            if "filled" in result or "(append)" in result:
                log("PASSED", result)
                changed += 1
                ok += 1
            else:
                log("INFO", result)
                ok += 1
        if ok == 0:
            raise RuntimeError("No SharePoint sheet was updated")
        if changed == 0:
            log(
                "PASSED",
                f"SharePoint already complete for {today.isoformat()} — leave workbook unchanged",
            )
        else:
            log("PASSED", "SharePoint Excel updated in-place (Graph) for station sheets")
        return 0
    finally:
        close_workbook_session(token, drive_id, item_id, session_id)


def acquire_graph_token(
    tenant: str,
    username: str,
    password: str,
    cfg_client_id: str,
    cache_path: str,
    site_url: str,
    allow_interactive: bool,
) -> str | None:
    """Prefer a Microsoft Graph token for Excel workbook APIs."""
    try_password = bool(password) and sys.platform != "win32"
    graph_scope_sets = [
        ["https://graph.microsoft.com/.default"],
        list(GRAPH_SCOPES),
        ["https://graph.microsoft.com/Files.ReadWrite.All"],
    ]
    for cid in client_ids_to_try(cfg_client_id):
        for scopes in graph_scope_sets:
            try:
                token = acquire_token_silent_or_password(
                    tenant, cid, username, password, cache_path, scopes, try_password
                )
                if token:
                    log("INFO", f"Graph token OK (client {cid[:8]}…)")
                    return token
            except Exception as exc:
                log("DEBUG", f"Graph silent/ROPC {cid[:8]}: {exc}")
    if allow_interactive:
        for cid in client_ids_to_try(cfg_client_id):
            for scopes in (
                ["https://graph.microsoft.com/.default"],
                ["https://graph.microsoft.com/Files.ReadWrite.All"],
            ):
                try:
                    return acquire_token_device_code(tenant, cid, cache_path, scopes)
                except Exception as exc:
                    log("DEBUG", f"Graph device-code {cid[:8]}: {exc}")
    return None


def make_sp_ctx_from_token(site_url: str, token: str):
    from office365.sharepoint.client_context import ClientContext

    def _token():
        return {"token_type": "Bearer", "access_token": token}

    ctx = ClientContext(site_url).with_access_token(_token)
    ctx.web.get().execute_query()
    log("INFO", f"SharePoint token OK — site '{ctx.web.properties.get('Title', site_url)}'")
    return ctx


def make_sp_ctx(site_url: str, tenant: str, client_id: str, username: str, password: str):
    from office365.sharepoint.client_context import ClientContext

    ctx = ClientContext(site_url).with_username_and_password(tenant, client_id, username, password)
    ctx.web.get().execute_query()
    log("INFO", f"SharePoint login OK — site '{ctx.web.properties.get('Title', site_url)}'")
    return ctx


def download_by_path(ctx, file_path: str) -> bytes:
    buf = io.BytesIO()
    try:
        ctx.web.get_file_by_server_relative_path(file_path).download(buf).execute_query()
    except Exception:
        ctx.web.get_file_by_server_relative_url(file_path).download(buf).execute_query()
    data = buf.getvalue()
    if not data:
        raise RuntimeError(f"Empty download for {file_path}")
    return data


def upload_by_path(ctx, file_path: str, data: bytes) -> None:
    folder = file_path.rsplit("/", 1)[0]
    name = file_path.rsplit("/", 1)[-1]
    ctx.web.get_folder_by_server_relative_url(folder).upload_file(name, data).execute_query()


def find_file_path(ctx, site_url: str, file_name: str) -> str | None:
    from office365.sharepoint.listitems.caml.query import CamlQuery

    leaf = os.path.basename(file_name)
    stem = os.path.splitext(leaf)[0]
    caml = f"""
    <View Scope='RecursiveAll'>
      <Query>
        <Where>
          <And>
            <Contains>
              <FieldRef Name='FileLeafRef'/>
              <Value Type='Text'>{stem}</Value>
            </Contains>
            <Eq>
              <FieldRef Name='FSObjType'/>
              <Value Type='Integer'>0</Value>
            </Eq>
          </And>
        </Where>
      </Query>
      <ViewFields>
        <FieldRef Name='FileRef'/>
        <FieldRef Name='FileLeafRef'/>
      </ViewFields>
      <RowLimit>10</RowLimit>
    </View>
    """
    try:
        lib = ctx.web.default_document_library()
        items = lib.get_items(CamlQuery.parse(caml)).execute_query()
        for item in items:
            ref = item.properties.get("FileRef") or ""
            if ref.lower().endswith(".xlsx"):
                log("INFO", f"Found workbook via search: {ref}")
                return ref
    except Exception as exc:
        log("DEBUG", f"CAML search failed: {exc}")

    parsed = urlparse(site_url)
    site_path = parsed.path.rstrip("/") or "/sites/simplifi-qa"
    candidates = [
        f"{site_path}/Shared Documents/{leaf}",
        f"{site_path}/Documents/{leaf}",
        f"{site_path}/Shared Documents/General/{leaf}",
    ]
    for rel in candidates:
        try:
            f = ctx.web.get_file_by_server_relative_url(rel)
            ctx.load(f)
            ctx.execute_query()
            if f.properties.get("Exists", True):
                log("INFO", f"Found workbook at {rel}")
                return rel
        except Exception:
            continue
    return None


def write_local_xlsx(path: str, data: bytes) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)
    try:
        with open(path, "wb") as f:
            f.write(data)
        saved = os.path.abspath(path)
    except PermissionError:
        root, ext = os.path.splitext(path)
        saved = os.path.abspath(f"{root}.new{ext or '.xlsx'}")
        with open(saved, "wb") as f:
            f.write(data)
        log("WARN", f"{path} is open in Excel — saved as {saved} instead")
    log("PASSED", f"Dry-run workbook saved: {saved} (SharePoint not changed)")


def apply_today_or_skip(
    raw: bytes,
    csv_rows: dict[str, dict[str, str]],
    today: date,
    dry_run: bool,
    out_xlsx: str,
    upload,
    sheets: list | None = None,
) -> int:
    """Fill new rows or blank cells (e.g. Uptime). Skip upload if nothing changed."""
    updated, changed = fill_workbook(raw, csv_rows, today, sheets=sheets)
    if changed == 0:
        log(
            "PASSED",
            f"SharePoint already complete for {today.isoformat()} — leave workbook unchanged",
        )
        return 0
    if dry_run:
        write_local_xlsx(out_xlsx, updated)
        return 0
    log("INFO", "Uploading workbook")
    upload(updated)
    log("PASSED", "SharePoint Excel updated for VN + US station sheets")
    return 0


def fill_workbook(
    data: bytes,
    csv_rows: dict[str, dict[str, str]],
    today: date,
    sheets: list | None = None,
) -> tuple[bytes, int]:
    wb = load_workbook(io.BytesIO(data))
    log("INFO", f"Workbook sheets: {', '.join(wb.sheetnames)}")
    ok = 0
    changed = 0
    for spec in sheets or SHEETS_FALLBACK:
        rec = csv_rows.get(spec["imei"])
        if not rec:
            log("WARN", f"No CSV row today for IMEI {spec['imei']} ({spec['names'][0]}) — skip sheet")
            continue
        result = upsert_sheet(wb, spec, rec, today)
        if result.startswith("sheet not found") or "missing" in result:
            log("ERROR", result)
            continue
        if "filled" in result or "(append)" in result:
            log("PASSED", result)
            changed += 1
            ok += 1
        else:
            log("INFO", result)
            ok += 1
    if ok == 0:
        raise RuntimeError("No SharePoint sheet was updated")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--conf", default="sharepoint.conf")
    parser.add_argument("--monitor-conf", default="monitor.conf")
    parser.add_argument("--csv", default="output/daily_monitor.csv")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Download + fill locally only (no SharePoint write). Normal runs edit cells in-place via Graph.",
    )
    parser.add_argument("--interactive", action="store_true",
                        help="If password login fails, open a Microsoft browser sign-in")
    parser.add_argument("--out-xlsx", default="output/monitoring_station.filled.xlsx")
    parser.add_argument("--date", default="", help="YYYY-MM-DD (default: today)")
    args = parser.parse_args()

    cfg = load_conf(args.conf)
    username = cfg.get("username") or cfg.get("email") or ""
    password = cfg.get("password") or ""
    site_url = (cfg.get("site_url") or "https://simplifirouter.sharepoint.com/sites/simplifi-qa").rstrip("/")
    file_url = cfg.get("file_url") or ""
    file_path = cfg.get("file_path") or ""
    file_name = cfg.get("file_name") or "monitoring_station.xlsx"
    csv_path = cfg.get("csv") or args.csv
    tenant = infer_tenant(username, cfg.get("tenant") or "")
    cfg_client_id = cfg.get("client_id") or ""
    cache_path = cfg.get("token_cache") or "token_cache.bin"
    if args.date:
        today = datetime.strptime(args.date.strip()[:10], "%Y-%m-%d").date()
    else:
        today = datetime.now().date()
    dry_run = args.dry_run
    out_xlsx = args.out_xlsx

    allow_interactive = args.interactive or sys.platform == "win32"
    if not username:
        log("ERROR", f"Fill username= in {args.conf}")
        return 2
    if not password and not allow_interactive:
        log("ERROR", f"Fill password= in {args.conf} (or pass --interactive)")
        return 2
    if not os.path.isfile(csv_path):
        log("ERROR", f"CSV not found: {csv_path}")
        return 2

    csv_rows = read_today_csv(csv_path, today)
    sheets = resolve_sheets(args.monitor_conf)
    wanted = [s["imei"] for s in sheets]
    have = [i for i in wanted if i in csv_rows]
    log("INFO", f"CSV date {today.isoformat()}: {len(have)}/{len(wanted)} lab stations (today only, not older rows)")
    if dry_run:
        log("INFO", "Dry-run: will not upload to SharePoint")
    if not have:
        log("ERROR", "No today's rows in CSV for the station IMEIs")
        return 1

    log("INFO", f"Microsoft login tenant={tenant}")
    token = None
    kind = "spo"
    try:
        token, kind = msal_login(
            tenant, username, password, cfg_client_id, cache_path, site_url,
            allow_interactive=allow_interactive,
        )
    except Exception as exc:
        log("ERROR", f"Microsoft login failed: {exc}")
        log("ERROR", "On this laptop: Cancel any 'Microsoft Graph Command Line Tools' approval page, "
            "delete token_cache.bin if it exists, then rerun. Sign in as Microsoft Office / Edge, "
            "not Graph CLI.")
        return 1

    graph_token = token if kind == "graph" else None
    if not graph_token:
        graph_token = acquire_graph_token(
            tenant, username, password, cfg_client_id, cache_path, site_url, allow_interactive
        )

    try:
        # Prefer in-place Graph Excel edits (no full-file PUT / 423 locks).
        if graph_token and file_url and not dry_run:
            resolved = resolve_share(graph_token, file_url, api_root_for("graph", site_url))
            if not resolved and token and kind == "spo":
                # Resolve via SPO shares, then edit with Graph token if drive/item ids work
                resolved = resolve_share(token, file_url, api_root_for("spo", site_url))
            if resolved:
                drive_id, item_id = resolved
                try:
                    log("INFO", "Updating workbook in-place via Microsoft Graph Excel API")
                    return fill_workbook_via_graph(
                        graph_token, drive_id, item_id, csv_rows, today, sheets=sheets
                    )
                except Exception as exc:
                    log("WARN", f"In-place Graph fill failed ({exc}); falling back to download/upload")

        if token and file_url:
            api_root = api_root_for(kind, site_url)
            resolved = resolve_share(token, file_url, api_root)
            if not resolved and kind == "spo":
                resolved = resolve_share(token, file_url, api_root_for("graph", site_url))
                if resolved:
                    api_root = api_root_for("graph", site_url)
            if resolved:
                drive_id, item_id = resolved
                log("INFO", f"Downloading workbook via {kind}")
                raw = download_drive_item(token, drive_id, item_id, api_root)
                log("INFO", f"Downloaded {len(raw)} bytes")
                return apply_today_or_skip(
                    raw, csv_rows, today, dry_run, out_xlsx,
                    lambda data: upload_drive_item(token, drive_id, item_id, api_root, data),
                    sheets=sheets,
                )

        last = None
        ctx = None
        if token:
            try:
                ctx = make_sp_ctx_from_token(site_url, token)
            except Exception as exc:
                last = exc
                log("DEBUG", f"SharePoint token context failed: {exc}")
        if ctx is None and password and sys.platform != "win32":
            for cid in client_ids_to_try(cfg_client_id):
                try:
                    ctx = make_sp_ctx(site_url, tenant, cid, username, password)
                    break
                except Exception as exc:
                    last = exc
                    log("DEBUG", f"SharePoint ROPC client {cid[:8]}… failed: {exc}")
        if ctx is None:
            raise last or RuntimeError("SharePoint login failed")

        rel_path = file_path
        if not rel_path:
            rel_path = find_file_path(ctx, site_url, file_name) or ""
        if not rel_path:
            log("ERROR", "Could not locate the monitoring Excel — set file_path= in sharepoint.conf")
            return 1
        log("INFO", "Downloading workbook via SharePoint")
        raw = download_by_path(ctx, rel_path)
        log("INFO", f"Downloaded {len(raw)} bytes")
        return apply_today_or_skip(
            raw, csv_rows, today, dry_run, out_xlsx,
            lambda data: upload_by_path(ctx, rel_path, data),
            sheets=sheets,
        )
    except Exception as exc:
        log("ERROR", f"{type(exc).__name__}: {exc}")
        msg = str(exc)
        if "AADSTS50076" in msg or "multi-factor" in msg.lower() or "MFA" in msg:
            log("ERROR", "This account requires MFA — ROPC cannot work. Use a non-MFA account or an Azure app + refresh token.")
        elif "AADSTS65001" in msg or "consent" in msg.lower():
            log("ERROR", "Azure admin consent is required for this app, or set client_id= of a consented public client in sharepoint.conf")
        return 1


if __name__ == "__main__":
    sys.exit(main())
